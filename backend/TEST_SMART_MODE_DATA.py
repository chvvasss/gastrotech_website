"""
Script to generate test Excel file for Smart Mode verification.

This creates an Excel file with the exact template structure and tests:
1. Smart mode candidate creation (missing Category/Series/Brand)
2. Product and Variant creation
3. Spec columns parsing
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill

# Test data that should trigger smart mode candidates
products_data = [
    {
        'Brand': 'GastroTech',  # Will create if doesn't exist
        'Category': 'Pişirme Üniteleri',  # Will create if doesn't exist
        'Series': '600 Series',  # Will create if doesn't exist
        'Product Name': 'Gazlı Ocak',
        'Product Slug': 'gazli-ocak-600',
        'Title TR': 'Endüstriyel Gazlı Ocak',
        'Title EN': 'Industrial Gas Stove',
        'Status': 'active',
        'Is Featured': 'yes',
        'Long Description': 'Profesyonel mutfaklar için yüksek kapasiteli gazlı ocak',
        'General Features': 'Paslanmaz çelik gövde, Kolay temizlik, Yüksek verim',
        'Short Specs': '6 gözlü, LPG/NG, 220V',
        'Taxonomy': '',
    },
    {
        'Brand': 'GastroTech',
        'Category': 'Pişirme Üniteleri',
        'Series': '700 Series',  # Another new series
        'Product Name': 'Elektrikli Fırın',
        'Product Slug': 'elektrikli-firin-700',
        'Title TR': 'Konveksiyonlu Elektrikli Fırın',
        'Title EN': 'Convection Electric Oven',
        'Status': 'active',
        'Is Featured': 'no',
        'Long Description': 'Konveksiyonlu pişirme sistemi ile homojen sonuçlar',
        'General Features': 'Dijital kontrol, Timer, Otomatik kapatma',
        'Short Specs': '5 tepsi, 380V, Konveksiyonlu',
        'Taxonomy': '',
    },
]

variants_data = [
    {
        'Product Slug': 'gazli-ocak-600',
        'Model Code': 'GO-600-6',
        'Variant Name TR': '6 Gözlü Model',
        'Variant Name EN': '6 Burner Model',
        'SKU': 'GO6006',
        'Dimensions': '600x700x280',
        'Weight': '85.5',
        'List Price': '15000.00',
        'Stock Qty': '10',
        'Spec:Power': '12 kW',
        'Spec:Capacity': '6 gözlü',
        'Spec:Material': 'Paslanmaz Çelik 304',
        'Spec:Voltage': '220V',
    },
    {
        'Product Slug': 'gazli-ocak-600',
        'Model Code': 'GO-600-8',
        'Variant Name TR': '8 Gözlü Model',
        'Variant Name EN': '8 Burner Model',
        'SKU': 'GO6008',
        'Dimensions': '800x700x280',
        'Weight': '95.0',
        'List Price': '18000.00',
        'Stock Qty': '5',
        'Spec:Power': '16 kW',
        'Spec:Capacity': '8 gözlü',
        'Spec:Material': 'Paslanmaz Çelik 304',
        'Spec:Voltage': '220V',
    },
    {
        'Product Slug': 'elektrikli-firin-700',
        'Model Code': 'EF-700-5',
        'Variant Name TR': '5 Tepsi Model',
        'Variant Name EN': '5 Tray Model',
        'SKU': 'EF7005',
        'Dimensions': '800x900x1800',
        'Weight': '120.0',
        'List Price': '35000.00',
        'Stock Qty': '3',
        'Spec:Power': '12 kW',
        'Spec:Capacity': '5 tepsi',
        'Spec:Material': 'Paslanmaz Çelik 304',
        'Spec:Voltage': '380V 3-Phase',
    },
]

# Create Excel file
output_file = 'test_smart_mode.xlsx'

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    # Products sheet
    df_products = pd.DataFrame(products_data)
    df_products.to_excel(writer, sheet_name='Products', index=False)

    # Variants sheet
    df_variants = pd.DataFrame(variants_data)
    df_variants.to_excel(writer, sheet_name='Variants', index=False)

    # Format headers
    workbook = writer.book
    for sheet_name in ['Products', 'Variants']:
        ws = workbook[sheet_name]
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

print(f"✅ Test file created: {output_file}")
print()
print("Test data summary:")
print(f"  - Products: {len(products_data)} (2 products)")
print(f"  - Variants: {len(variants_data)} (3 variants)")
print()
print("Smart mode candidates (if entities don't exist):")
print("  - Brand: GastroTech (slug: gastrotech)")
print("  - Category: Pişirme Üniteleri (slug: pisirme-uniteleri)")
print("  - Series: 600 Series (slug: 600-series), 700 Series (slug: 700-series)")
print()
print("Next steps:")
print("1. Upload this file via /api/admin/import-jobs/validate/ with mode=smart")
print("2. Check response.data.candidates for detected missing entities")
print("3. Commit via /api/admin/import-jobs/{job_id}/commit/")
print("4. Verify DB: Category.objects.filter(slug='pisirme-uniteleri').exists()")
