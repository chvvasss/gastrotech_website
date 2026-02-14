"""
Compare Excel products with database products and generate a report.
"""
import os
import sys
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings.dev')
django.setup()

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from apps.catalog.models import Variant, Product

# Excel'deki model kodlarını oku
wb_source = openpyxl.load_workbook(r'C:\gastrotech.com.tr.0101\gastrotech.com_cursor\Yeni klasör\Gastrotech_Tum_Veriler_Duzenlenmis.xlsx')
ws_source = wb_source['Urunler']

excel_products = []
for row in range(2, ws_source.max_row + 1):
    model_code = ws_source.cell(row=row, column=4).value
    urun_adi_tr = ws_source.cell(row=row, column=2).value
    urun_adi_en = ws_source.cell(row=row, column=3).value
    seri = ws_source.cell(row=row, column=5).value
    kategori = ws_source.cell(row=row, column=6).value
    alt_kategori = ws_source.cell(row=row, column=7).value

    if model_code and str(model_code).strip() != '-':
        excel_products.append({
            'model_code': str(model_code).strip(),
            'urun_adi_tr': urun_adi_tr,
            'urun_adi_en': urun_adi_en,
            'seri': seri,
            'kategori': kategori,
            'alt_kategori': alt_kategori,
        })

excel_model_codes = set(p['model_code'] for p in excel_products)

# Sitedeki model kodları
db_variants = list(Variant.objects.select_related('product', 'product__series', 'product__category', 'product__series__category').all())
db_model_codes = set(v.model_code for v in db_variants)

# Karşılaştırma
only_in_excel = excel_model_codes - db_model_codes
only_in_db = db_model_codes - excel_model_codes
in_both = excel_model_codes & db_model_codes

# Rapor Excel'i oluştur
wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, color="FFFFFF")
header_fill_red = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
header_fill_green = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
header_fill_blue = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# ===== SAYFA 1: ÖZET =====
ws_summary = wb.active
ws_summary.title = "Ozet"

ws_summary.cell(row=1, column=1, value="KARSILASTIRMA OZETI").font = Font(bold=True, size=14)
ws_summary.cell(row=3, column=1, value="Metrik")
ws_summary.cell(row=3, column=2, value="Deger")

summary_data = [
    ("Excel dosyasinda toplam model kodu", len(excel_model_codes)),
    ("Sitede (DB) toplam model kodu", len(db_model_codes)),
    ("", ""),
    ("Her ikisinde de olan", len(in_both)),
    ("SADECE Excel'de olan (sitede YOK)", len(only_in_excel)),
    ("SADECE sitede olan (Excel'de YOK)", len(only_in_db)),
]

for i, (metric, value) in enumerate(summary_data, 4):
    ws_summary.cell(row=i, column=1, value=metric)
    ws_summary.cell(row=i, column=2, value=value)

ws_summary.column_dimensions['A'].width = 40
ws_summary.column_dimensions['B'].width = 15

# ===== SAYFA 2: Excel'de olup sitede olmayan =====
ws_missing = wb.create_sheet("Sitede_Eksik_Urunler")

headers = ["Model Kodu", "Urun Adi TR", "Urun Adi EN", "Seri", "Kategori", "Alt Kategori"]
for col, header in enumerate(headers, 1):
    cell = ws_missing.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill_red
    cell.border = thin_border

excel_only_products = sorted([p for p in excel_products if p['model_code'] in only_in_excel],
                              key=lambda x: (x['kategori'] or '', x['model_code']))

for row_num, p in enumerate(excel_only_products, 2):
    ws_missing.cell(row=row_num, column=1, value=p['model_code']).border = thin_border
    ws_missing.cell(row=row_num, column=2, value=p['urun_adi_tr']).border = thin_border
    ws_missing.cell(row=row_num, column=3, value=p['urun_adi_en']).border = thin_border
    ws_missing.cell(row=row_num, column=4, value=p['seri']).border = thin_border
    ws_missing.cell(row=row_num, column=5, value=p['kategori']).border = thin_border
    ws_missing.cell(row=row_num, column=6, value=p['alt_kategori']).border = thin_border

ws_missing.column_dimensions['A'].width = 25
ws_missing.column_dimensions['B'].width = 50
ws_missing.column_dimensions['C'].width = 50
ws_missing.column_dimensions['D'].width = 25
ws_missing.column_dimensions['E'].width = 25
ws_missing.column_dimensions['F'].width = 25
ws_missing.freeze_panes = "A2"

# ===== SAYFA 3: Sitede olup Excel'de olmayan =====
ws_extra = wb.create_sheet("Excelde_Olmayan_Urunler")

headers = ["Model Kodu", "Urun Adi TR", "Urun Adi EN", "Seri", "Kategori"]
for col, header in enumerate(headers, 1):
    cell = ws_extra.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill_green
    cell.border = thin_border

db_only_variants = sorted([v for v in db_variants if v.model_code in only_in_db],
                           key=lambda x: (x.product.series.name if x.product and x.product.series else '', x.model_code))

for row_num, v in enumerate(db_only_variants, 2):
    series_name = v.product.series.name if v.product and v.product.series else ''
    cat_name = ''
    if v.product and v.product.series and v.product.series.category:
        cat_name = v.product.series.category.name

    ws_extra.cell(row=row_num, column=1, value=v.model_code).border = thin_border
    ws_extra.cell(row=row_num, column=2, value=v.product.title_tr if v.product else '').border = thin_border
    ws_extra.cell(row=row_num, column=3, value=v.product.title_en if v.product else '').border = thin_border
    ws_extra.cell(row=row_num, column=4, value=series_name).border = thin_border
    ws_extra.cell(row=row_num, column=5, value=cat_name).border = thin_border

ws_extra.column_dimensions['A'].width = 25
ws_extra.column_dimensions['B'].width = 50
ws_extra.column_dimensions['C'].width = 50
ws_extra.column_dimensions['D'].width = 25
ws_extra.column_dimensions['E'].width = 25
ws_extra.freeze_panes = "A2"

# ===== SAYFA 4: Kategori bazlı özet =====
ws_cat = wb.create_sheet("Kategori_Bazli_Eksikler")

headers = ["Kategori", "Excel'de Olan", "Sitede Eksik", "Eksik Orani (%)"]
for col, header in enumerate(headers, 1):
    cell = ws_cat.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill_blue
    cell.border = thin_border

# Kategoriye göre grupla
by_category = {}
for p in excel_products:
    cat = p['kategori'] or 'Kategorisiz'
    if cat not in by_category:
        by_category[cat] = {'total': 0, 'missing': 0}
    by_category[cat]['total'] += 1
    if p['model_code'] in only_in_excel:
        by_category[cat]['missing'] += 1

row_num = 2
for cat in sorted(by_category.keys()):
    data = by_category[cat]
    ratio = (data['missing'] / data['total'] * 100) if data['total'] > 0 else 0

    ws_cat.cell(row=row_num, column=1, value=cat).border = thin_border
    ws_cat.cell(row=row_num, column=2, value=data['total']).border = thin_border
    ws_cat.cell(row=row_num, column=3, value=data['missing']).border = thin_border
    ws_cat.cell(row=row_num, column=4, value=f"{ratio:.1f}%").border = thin_border
    row_num += 1

ws_cat.column_dimensions['A'].width = 35
ws_cat.column_dimensions['B'].width = 15
ws_cat.column_dimensions['C'].width = 15
ws_cat.column_dimensions['D'].width = 15
ws_cat.freeze_panes = "A2"

# Kaydet
output_file = "urun_karsilastirma_raporu.xlsx"
wb.save(output_file)

print("=" * 60)
print("KARSILASTIRMA RAPORU OLUSTURULDU")
print("=" * 60)
print(f"Dosya: {output_file}")
print()
print("OZET:")
print(f"  - Excel'de toplam: {len(excel_model_codes)} urun")
print(f"  - Sitede toplam: {len(db_model_codes)} urun")
print(f"  - Her ikisinde: {len(in_both)} urun")
print(f"  - SITEDE EKSIK: {len(only_in_excel)} urun")
print(f"  - EXCEL'DE YOK: {len(only_in_db)} urun")
print()
print("Rapor sayfalari:")
print("  1. Ozet - Genel istatistikler")
print("  2. Sitede_Eksik_Urunler - Excel'de olup sitede olmayan urunler")
print("  3. Excelde_Olmayan_Urunler - Sitede olup Excel'de olmayan urunler")
print("  4. Kategori_Bazli_Eksikler - Kategori bazinda eksik analizi")
