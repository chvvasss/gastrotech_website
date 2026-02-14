"""
Import remaining missing products (handle duplicate series).
"""
import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings.dev')
django.setup()

import openpyxl
from django.db import transaction
from apps.catalog.models import Category, Series, Product, Variant
from apps.common.slugify_tr import slugify_tr

# Mapping tablosu (daha önce tanımlananlarla aynı)
CATEGORY_MAPPING = {
    'FIRINLAR': 'Fırınlar',
    'Fırınlar': 'Fırınlar',
    'HAZIRLIK': 'Hazırlık Ekipmanları',
    'Hazırlık Ekipmanları': 'Hazırlık Ekipmanları',
    'KAFETERYA': 'Kafeterya Ekipmanları',
    'Kafeterya Ekipmanları': 'Kafeterya Ekipmanları',
    'Pişirme Ekipmanları': 'Pişirme Ekipmanları',
    'SOĞUTMA': 'Soğutma Üniteleri',
    'Soğutma Üniteleri': 'Soğutma Üniteleri',
    'Soğutma Ekipmanları': 'Soğutma Ekipmanları',
    'Tamamlayıcı Ekipmanlar': 'Tamamlayıcı Ekipmanlar',
    'Buz Makineleri': 'Buz Makineleri',
}

SERIES_MAPPING = {
    ('FIRINLAR', 'Taş Tabanlı'): ('Fırınlar', 'Taş Tabanlı Bakery Fırınlar'),
    ('FIRINLAR', 'KWIK-CO'): ('Fırınlar', 'KWIK-CO Serisi Konveksiyonel Fırınlar'),
    ('FIRINLAR', 'PRIME'): ('Fırınlar', 'PRIME'),
    ('FIRINLAR', 'NEVO'): ('Fırınlar', 'NEVO'),
    ('FIRINLAR', 'GR'): ('Fırınlar', 'GR'),
    ('FIRINLAR', 'MIX'): ('Fırınlar', 'MIX'),
    ('FIRINLAR', '-'): ('Fırınlar', 'Diğer Fırınlar'),
    ('Fırınlar', 'MAESTRO Serisi'): ('Fırınlar', 'MAESTRO Serisi'),
    ('Pişirme Ekipmanları', '600 Serisi'): ('Pişirme Ekipmanları', '600 Serisi'),
    ('Pişirme Ekipmanları', '700 Serisi'): ('Pişirme Ekipmanları', '700 Serisi'),
    ('Pişirme Ekipmanları', '900 Serisi'): ('Pişirme Ekipmanları', '900 Serisi'),
    ('Pişirme Ekipmanları', 'Drop-in Seri'): ('Pişirme Ekipmanları', 'Drop-in Serisi'),
    ('Pişirme Ekipmanları', 'Diğer'): ('Pişirme Ekipmanları', 'Diğer'),
    ('SOĞUTMA', 'Scotsman'): ('Buz Makineleri', 'Scotsman'),
    ('SOĞUTMA', 'Scotsman AC'): ('Buz Makineleri', 'AC Serisi'),
    ('SOĞUTMA', 'Scotsman AF'): ('Buz Makineleri', 'AF Serisi'),
    ('SOĞUTMA', 'Scotsman EC'): ('Buz Makineleri', 'EC Serisi'),
    ('SOĞUTMA', 'Scotsman MF'): ('Buz Makineleri', 'MF Serisi'),
    ('SOĞUTMA', 'Scotsman MXG'): ('Buz Makineleri', 'MXG Serisi'),
    ('SOĞUTMA', 'Scotsman NU'): ('Buz Makineleri', 'NU Serisi'),
    ('SOĞUTMA', 'Scotsman NW'): ('Buz Makineleri', 'NW Serisi'),
    ('SOĞUTMA', 'Scotsman SB'): ('Buz Makineleri', 'SB Serisi'),
    ('SOĞUTMA', 'Scotsman UBH'): ('Buz Makineleri', 'UBH Serisi'),
    ('SOĞUTMA', 'Scotsman LEGACY'): ('Buz Makineleri', 'LEGACY Serisi'),
    ('SOĞUTMA', 'Basic'): ('Soğutma Üniteleri', 'Basic Serisi'),
    ('SOĞUTMA', 'Premium'): ('Soğutma Üniteleri', 'Premium Serisi'),
    ('SOĞUTMA', 'B Serisi'): ('Soğutma Üniteleri', 'B Serisi'),
    ('SOĞUTMA', 'DKP Kasa'): ('Soğutma Üniteleri', 'DKP Kasa Serisi'),
    ('SOĞUTMA', 'DXN'): ('Soğutma Üniteleri', 'DXN Serisi'),
    ('SOĞUTMA', 'Paslanmaz Çelik'): ('Soğutma Üniteleri', 'Paslanmaz Çelik Serisi'),
    ('SOĞUTMA', 'Yerli Üretim'): ('Soğutma Üniteleri', 'Yerli Üretim'),
    ('SOĞUTMA', '-'): ('Soğutma Üniteleri', 'Diğer'),
    ('HAZIRLIK', 'Kitchen Aid'): ('Hazırlık Ekipmanları', 'Kitchen Aid'),
    ('HAZIRLIK', '-'): ('Hazırlık Ekipmanları', 'Diğer'),
    ('KAFETERYA', '-'): ('Kafeterya Ekipmanları', 'Diğer'),
    ('Kafeterya Ekipmanları', '-'): ('Kafeterya Ekipmanları', 'Diğer'),
    ('Kafeterya Ekipmanları', 'ALL GROUND'): ('Kafeterya Ekipmanları', 'ALL GROUND'),
    ('Tamamlayıcı Ekipmanlar', 'CBU Serisi'): ('Tamamlayıcı Ekipmanlar', 'CBU Serisi'),
    ('Tamamlayıcı Ekipmanlar', '-'): ('Tamamlayıcı Ekipmanlar', 'Diğer'),
}


def get_or_create_category(name):
    """Get or create category by name."""
    try:
        return Category.objects.get(name=name)
    except Category.DoesNotExist:
        slug = slugify_tr(name)
        cat = Category.objects.create(name=name, slug=slug)
        print(f"  [+] Kategori oluşturuldu: {name}")
        return cat


def get_or_create_series(category, series_name):
    """Get or create series by name in category - handle duplicates."""
    from django.db.models import Count

    # Önce mevcut serilere bak
    existing = Series.objects.filter(category=category, name=series_name)

    if existing.count() == 1:
        return existing.first()
    elif existing.count() > 1:
        # Duplicate var, en çok ürünü olan veya ilk oluşturulanı al
        series_with_products = existing.annotate(
            num_products=Count('products')
        ).order_by('-num_products', 'created_at')
        return series_with_products.first()
    else:
        # Yeni seri oluştur
        slug = slugify_tr(series_name)
        base_slug = slug
        counter = 1
        while Series.objects.filter(category=category, slug=slug).exists():
            slug = f"{base_slug}-{counter}"
            counter += 1

        series = Series.objects.create(
            category=category,
            name=series_name,
            slug=slug
        )
        print(f"  [+] Seri oluşturuldu: {category.name} > {series_name}")
        return series


def get_mapped_category_series(excel_category, excel_series):
    """Get mapped category and series names."""
    key = (excel_category, excel_series)
    if key in SERIES_MAPPING:
        return SERIES_MAPPING[key]

    mapped_category = CATEGORY_MAPPING.get(excel_category, excel_category)
    mapped_series = excel_series if excel_series and excel_series != '-' else 'Diğer'

    return (mapped_category, mapped_series)


def main():
    import django.db.models

    # Excel'deki verileri oku
    wb = openpyxl.load_workbook(
        r'C:\gastrotech.com.tr.0101\gastrotech.com_cursor\Yeni klasör\Gastrotech_Tum_Veriler_Duzenlenmis.xlsx'
    )
    ws = wb['Urunler']

    # Sitedeki mevcut model kodları
    existing_codes = set(Variant.objects.values_list('model_code', flat=True))
    print(f"Mevcut model kodu sayısı: {len(existing_codes)}")

    # Eksik ürünleri topla
    missing_products = []
    for row in range(2, ws.max_row + 1):
        model_code = ws.cell(row=row, column=4).value
        if model_code and str(model_code).strip() != '-':
            model_code = str(model_code).strip()
            if model_code not in existing_codes:
                missing_products.append({
                    'model_code': model_code,
                    'urun_adi_tr': ws.cell(row=row, column=2).value or model_code,
                    'urun_adi_en': ws.cell(row=row, column=3).value or '',
                    'seri': ws.cell(row=row, column=5).value or '-',
                    'kategori': ws.cell(row=row, column=6).value or 'Kategorisiz',
                    'alt_kategori': ws.cell(row=row, column=7).value or '',
                })

    print(f"Kalan eksik ürün: {len(missing_products)}")

    if not missing_products:
        print("Tüm ürünler zaten eklenmiş!")
        return

    print("=" * 60)

    # Ürünleri grupla
    product_groups = {}
    for p in missing_products:
        key = (p['urun_adi_tr'], p['kategori'], p['seri'])
        if key not in product_groups:
            product_groups[key] = {
                'urun_adi_tr': p['urun_adi_tr'],
                'urun_adi_en': p['urun_adi_en'],
                'kategori': p['kategori'],
                'seri': p['seri'],
                'variants': []
            }
        product_groups[key]['variants'].append(p['model_code'])

    print(f"Kalan ürün grubu: {len(product_groups)}")
    print("=" * 60)

    # Import işlemi
    created_products = 0
    created_variants = 0
    errors = []

    with transaction.atomic():
        for key, group in product_groups.items():
            try:
                mapped_cat, mapped_series = get_mapped_category_series(
                    group['kategori'], group['seri']
                )

                category = get_or_create_category(mapped_cat)
                series = get_or_create_series(category, mapped_series)

                product_name = group['urun_adi_tr']
                slug_base = slugify_tr(product_name)
                slug = slug_base
                counter = 1
                while Product.objects.filter(slug=slug).exists():
                    slug = f"{slug_base}-{counter}"
                    counter += 1

                product = Product.objects.create(
                    series=series,
                    category=category,
                    name=product_name,
                    slug=slug,
                    title_tr=product_name,
                    title_en=group['urun_adi_en'] or '',
                    status='active',
                )
                created_products += 1

                for model_code in group['variants']:
                    if Variant.objects.filter(model_code=model_code).exists():
                        continue

                    Variant.objects.create(
                        product=product,
                        model_code=model_code,
                        name_tr=product_name,
                        name_en=group['urun_adi_en'] or '',
                    )
                    created_variants += 1

                print(f"  [OK] {product_name} ({len(group['variants'])} varyant)")

            except Exception as e:
                errors.append(f"{group['urun_adi_tr']}: {str(e)}")
                print(f"  [ERROR] {group['urun_adi_tr']}: {str(e)}")

    print()
    print("=" * 60)
    print("SONUÇ")
    print("=" * 60)
    print(f"Oluşturulan ürün grubu: {created_products}")
    print(f"Oluşturulan varyant: {created_variants}")
    print(f"Hata sayısı: {len(errors)}")

    if errors:
        print("\nHatalar:")
        for err in errors[:20]:
            print(f"  - {err}")


if __name__ == '__main__':
    main()
