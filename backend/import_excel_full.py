"""
Full Excel import script for Gastrotech_Tum_Veriler.xlsx
Reads all 1173 rows, creates categories/series/brands/products/variants.
"""
import os, sys, django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings.dev')
sys.path.insert(0, '/app')
django.setup()

import openpyxl
from django.db import transaction
from django.utils.text import slugify
from apps.catalog.models import Category, Series, Brand, Product, Variant

# --- Slug mapping for categories ---
CAT_SLUG_MAP = {
    'Pişirme Ekipmanları': 'pisirme-ekipmanlari',
    'FIRINLAR': 'firinlar',
    'Fırınlar': 'firinlar',
    'SOĞUTMA': 'sogutma-uniteleri',
    'HAZIRLIK': 'hazirlik-ekipmanlari',
    'KAFETERYA': 'kafeterya-ekipmanlari',
    'Kafeterya Ekipmanları': 'kafeterya-ekipmanlari',
    'Bulaşık Makineleri': 'bulasikane',
    'Çamaşır Ekipmanları': 'camasirhane',
    'Tamamlayıcı Ekipmanlar': 'tamamlayici',
    'Aksesuarlar': 'aksesuarlar',
}

# Brand defaults
DEFAULT_BRAND_SLUG = 'gastrotech'

def get_or_create_category(name, parent=None):
    slug = CAT_SLUG_MAP.get(name)
    if not slug:
        slug = slugify(name, allow_unicode=False)
        if not slug:
            slug = slugify(name.replace('ı','i').replace('ö','o').replace('ü','u').replace('ş','s').replace('ç','c').replace('ğ','g'), allow_unicode=False)

    cat = Category.objects.filter(slug=slug).first()
    if cat:
        return cat

    # Try by name
    cat = Category.objects.filter(name__iexact=name).first()
    if cat:
        return cat

    cat = Category.objects.create(
        name=name,
        slug=slug,
        parent=parent,
        order=Category.objects.filter(parent=parent).count() + 1,
    )
    print(f'  [+] Category created: {name} ({slug})')
    return cat


def get_or_create_subcategory(name, parent_cat):
    slug = slugify(name, allow_unicode=False)
    if not slug:
        slug = slugify(name.replace('ı','i').replace('ö','o').replace('ü','u').replace('ş','s').replace('ç','c').replace('ğ','g'), allow_unicode=False)

    cat = Category.objects.filter(slug=slug, parent=parent_cat).first()
    if cat:
        return cat
    cat = Category.objects.filter(name__iexact=name, parent=parent_cat).first()
    if cat:
        return cat
    # Check without parent too
    cat = Category.objects.filter(slug=slug).first()
    if cat:
        return cat

    cat = Category.objects.create(
        name=name,
        slug=slug,
        parent=parent_cat,
        order=Category.objects.filter(parent=parent_cat).count() + 1,
    )
    print(f'  [+] Subcategory created: {name} under {parent_cat.name}')
    return cat


def get_or_create_series(name, category):
    if not name or name == '-':
        # Create a default "Genel" series for this category
        default_name = 'Genel'
        default_slug = f"{category.slug}-genel"
        s = Series.objects.filter(slug=default_slug).first()
        if s:
            return s
        s = Series.objects.create(
            name=default_name,
            slug=default_slug,
            category=category,
            order=999,
        )
        print(f'  [+] Default series created: {default_slug} in {category.name}')
        return s
    slug = slugify(name, allow_unicode=False)
    if not slug:
        slug = slugify(name.replace('ı','i').replace('ö','o').replace('ü','u').replace('ş','s').replace('ç','c').replace('ğ','g'), allow_unicode=False)

    s = Series.objects.filter(slug=slug).first()
    if s:
        return s
    s = Series.objects.filter(name__iexact=name, category=category).first()
    if s:
        return s

    s = Series.objects.create(
        name=name,
        slug=slug,
        category=category,
        order=Series.objects.filter(category=category).count() + 1,
    )
    print(f'  [+] Series created: {name} ({slug}) in {category.name}')
    return s


def get_or_create_brand(name):
    slug = slugify(name, allow_unicode=False)
    if not slug:
        slug = slugify(name.replace('ı','i').replace('ö','o').replace('ü','u').replace('ş','s').replace('ç','c').replace('ğ','g'), allow_unicode=False)

    b = Brand.objects.filter(slug=slug).first()
    if b:
        return b
    b = Brand.objects.filter(name__iexact=name).first()
    if b:
        return b

    b = Brand.objects.create(name=name, slug=slug, is_active=True, order=0)
    print(f'  [+] Brand created: {name}')
    return b


def run():
    wb = openpyxl.load_workbook('/app/Gastrotech_Tum_Veriler.xlsx')
    ws = wb['Urunler']

    default_brand = Brand.objects.filter(slug=DEFAULT_BRAND_SLUG).first()
    if not default_brand:
        default_brand = Brand.objects.filter(slug='gtech').first()
    if not default_brand:
        default_brand = get_or_create_brand('Gastrotech')

    stats = {
        'rows': 0, 'products_created': 0, 'products_updated': 0,
        'variants_created': 0, 'variants_updated': 0,
        'skipped': 0, 'errors': []
    }

    # Group rows by product (same Urun Adi TR + Seri + Kategori = same product, different model codes = variants)
    # Actually, each row is a variant with model code. Products are grouped by name+series+category.
    product_groups = {}

    for r in range(2, ws.max_row + 1):
        urun_adi_tr = ws.cell(r, 2).value
        urun_adi_en = ws.cell(r, 3).value
        model_code = ws.cell(r, 4).value
        seri = ws.cell(r, 5).value
        kategori = ws.cell(r, 6).value
        alt_kategori = ws.cell(r, 7).value

        if not urun_adi_tr or not model_code:
            continue

        model_code = str(model_code).strip()
        if model_code == '-' or not model_code:
            continue

        # Product grouping key
        key = (str(urun_adi_tr).strip(), str(seri or '').strip(), str(kategori or '').strip())

        if key not in product_groups:
            product_groups[key] = {
                'name_tr': str(urun_adi_tr).strip(),
                'name_en': str(urun_adi_en or '').strip(),
                'series': str(seri or '').strip(),
                'category': str(kategori or '').strip(),
                'subcategory': str(alt_kategori or '').strip(),
                'variants': []
            }

        product_groups[key]['variants'].append({
            'model_code': model_code,
            'name_tr': str(urun_adi_tr).strip(),
            'name_en': str(urun_adi_en or '').strip(),
        })
        stats['rows'] += 1

    print(f'\nParsed {stats["rows"]} rows into {len(product_groups)} product groups')
    print('Starting import...\n')

    with transaction.atomic():
        for key, group in product_groups.items():
            try:
                # Get/create category
                cat_name = group['category']
                if not cat_name:
                    stats['skipped'] += 1
                    continue

                parent_cat = get_or_create_category(cat_name)

                # Get/create subcategory
                target_cat = parent_cat
                if group['subcategory'] and group['subcategory'] != 'None':
                    target_cat = get_or_create_subcategory(group['subcategory'], parent_cat)

                # Get/create series
                series = get_or_create_series(group['series'], target_cat)

                # Build product slug
                base_slug = slugify(group['name_tr'], allow_unicode=False)
                if not base_slug:
                    base_slug = slugify(group['name_tr'].replace('ı','i').replace('ö','o').replace('ü','u').replace('ş','s').replace('ç','c').replace('ğ','g'), allow_unicode=False)

                if series:
                    series_slug = series.slug
                    product_slug = f"{base_slug}-{series_slug}" if base_slug else series_slug
                else:
                    product_slug = base_slug

                if not product_slug:
                    product_slug = slugify(group['variants'][0]['model_code'])

                # Ensure unique slug
                orig_slug = product_slug
                counter = 1
                while Product.objects.filter(slug=product_slug).exclude(
                    name=group['name_tr'], series=series, category=target_cat
                ).exists():
                    product_slug = f"{orig_slug}-{counter}"
                    counter += 1

                # Upsert product
                product, created = Product.objects.update_or_create(
                    slug=product_slug,
                    defaults={
                        'name': group['name_tr'],
                        'title_tr': group['name_tr'],
                        'title_en': group['name_en'],
                        'status': 'active',
                        'series': series,
                        'category': target_cat,
                        'brand': default_brand,
                    }
                )

                if created:
                    stats['products_created'] += 1
                else:
                    stats['products_updated'] += 1

                # Create variants
                for v in group['variants']:
                    mc = v['model_code']
                    variant, v_created = Variant.objects.update_or_create(
                        model_code=mc,
                        defaults={
                            'product': product,
                            'name_tr': v['name_tr'],
                            'name_en': v['name_en'],
                        }
                    )
                    if v_created:
                        stats['variants_created'] += 1
                    else:
                        stats['variants_updated'] += 1

            except Exception as e:
                err = f"Error for {key}: {e}"
                stats['errors'].append(err)
                print(f'  [!] {err}')

    print('\n' + '=' * 60)
    print('IMPORT SUMMARY')
    print('=' * 60)
    print(f'Total rows processed:  {stats["rows"]}')
    print(f'Product groups:        {len(product_groups)}')
    print(f'Products created:      {stats["products_created"]}')
    print(f'Products updated:      {stats["products_updated"]}')
    print(f'Variants created:      {stats["variants_created"]}')
    print(f'Variants updated:      {stats["variants_updated"]}')
    print(f'Skipped:               {stats["skipped"]}')
    print(f'Errors:                {len(stats["errors"])}')
    if stats['errors']:
        print('\nFirst 10 errors:')
        for e in stats['errors'][:10]:
            print(f'  - {e}')


if __name__ == '__main__':
    run()
