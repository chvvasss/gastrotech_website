"""Create product upload Excel template."""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
section_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
required_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
wrap = Alignment(wrap_text=True, vertical='top')

# ===== SAYFA 1: URUN YUKLEME SABLONU =====
ws = wb.active
ws.title = 'Urun_Yukleme'

columns = [
    # (Header, Width, Required, Description)
    ('Sira No', 8, False, 'Otomatik sira'),
    ('Kategori', 25, True, 'Ana kategori adi'),
    ('Seri', 25, True, 'Seri adi'),
    ('Urun Adi TR', 40, True, 'Turkce urun adi'),
    ('Urun Adi EN', 40, False, 'Ingilizce urun adi'),
    ('Model Kodu', 20, True, 'Benzersiz model kodu'),
    ('SKU', 15, False, 'Stok birimi kodu'),
    ('Boyutlar (GxDxY mm)', 22, False, 'Genislik x Derinlik x Yukseklik'),
    ('Agirlik (kg)', 12, False, 'Net agirlik'),
    ('Guc (W/kW)', 12, False, 'Elektrik gucu'),
    ('Voltaj', 28, False, 'Ornek: 220-230V 1N AC 50Hz'),
    ('Kapasite', 15, False, 'Litre, kisi, GN vb.'),
    ('Yakit Tipi', 15, False, 'Elektrik / Dogalgaz / LPG'),
    ('Malzeme', 22, False, 'Paslanmaz celik, pik dokum vb.'),
    ('Ek Spec 1 Anahtar', 18, False, 'Ornek: pilot_flame'),
    ('Ek Spec 1 Deger', 18, False, 'Ornek: evet'),
    ('Ek Spec 2 Anahtar', 18, False, 'Ornek: thermostat'),
    ('Ek Spec 2 Deger', 18, False, 'Ornek: 50-300C'),
    ('Ek Spec 3 Anahtar', 18, False, 'Serbest'),
    ('Ek Spec 3 Deger', 18, False, 'Serbest'),
    ('Genel Ozellikler TR', 50, False, 'Satirlar | ile ayrilir'),
    ('Genel Ozellikler EN', 50, False, 'Lines separated by |'),
    ('Uzun Aciklama TR', 50, False, 'Detayli aciklama'),
    ('Uzun Aciklama EN', 50, False, 'Detailed description'),
    ('Liste Fiyati', 15, False, 'Fiyat'),
    ('Durum', 12, False, 'active / draft / archived'),
    ('Foto 1 (Sayfa-Sira)', 18, True, 'PDF sayfa no - gorsel sirasi (ornek: 9-1)'),
    ('Foto 2 (Sayfa-Sira)', 18, False, 'ornek: 9-2'),
    ('Foto 3 (Sayfa-Sira)', 18, False, 'ornek: 9-3'),
    ('Foto 4 (Sayfa-Sira)', 18, False, 'ornek: 10-1'),
    ('Foto 5 (Sayfa-Sira)', 18, False, 'ornek: 10-2'),
    ('Notlar', 30, False, 'Ek bilgiler'),
]

# Write headers
for col, (name, width, required, desc) in enumerate(columns, 1):
    cell = ws.cell(row=1, column=col, value=name)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    ws.column_dimensions[get_column_letter(col)].width = width

# Write description row
for col, (name, width, required, desc) in enumerate(columns, 1):
    cell = ws.cell(row=2, column=col, value=desc)
    cell.fill = section_fill
    cell.border = thin_border
    cell.alignment = wrap
    cell.font = Font(size=9, italic=True)

# Mark required columns yellow for data rows
for col, (name, width, required, desc) in enumerate(columns, 1):
    if required:
        for row in range(3, 203):
            ws.cell(row=row, column=col).fill = required_fill
            ws.cell(row=row, column=col).border = thin_border

# Data validation for Durum
dv = DataValidation(type='list', formula1='"active,draft,archived"', allow_blank=True)
ws.add_data_validation(dv)
durum_col = 26
for row in range(3, 203):
    dv.add(ws.cell(row=row, column=durum_col))

# Sample data rows
samples = [
    [1, 'Pisirme Ekipmanlari', '600 Serisi', '2 Yanisli Gazli Ocak',
     '2 Burner Gas Range', 'GKO6010', '', '400x700x300', 35,
     12, '220-230V 1N AC 50Hz', '', 'Dogalgaz',
     'Paslanmaz celik', 'pilot_flame', 'evet', 'thermocouple', 'evet', '', '',
     'Paslanmaz celik govde|Pres baski yekpare ust tabla|Gaz emniyet tertibati ve termokupl|Pik dokum cikarilabilir izgaralar',
     'Stainless steel body|Press-formed one-piece top|Gas safety device with thermocouple|Cast iron removable grids',
     '', '', '', 'active',
     '9-1', '9-2', '', '', '', ''],
    [2, 'Pisirme Ekipmanlari', '600 Serisi', '2 Yanisli Gazli Ocak',
     '2 Burner Gas Range', 'GKW6010', '', '400x700x300', 35,
     12, '220-230V 1N AC 50Hz', '', 'Dogalgaz',
     'Paslanmaz celik', 'wok_burner', 'evet', '', '', '', '',
     '', '', '', '', '', 'active',
     '9-3', '9-4', '', '', '', 'Wok varyanti'],
    [3, 'Pisirme Ekipmanlari', '600 Serisi', '4 Yanisli Gazli Ocak Firinli',
     '4 Burner Gas Range w/ Oven', 'GKF6020', '', '800x700x850', 85,
     24, '220-230V 1N AC 50Hz', '', 'Dogalgaz',
     'Paslanmaz celik', 'oven_type', 'statik', 'oven_temp', '50-300C', '', '',
     'Paslanmaz celik govde|4 yanisli gazli ocak|Statik firinli|Termostatik kumanda',
     'Stainless steel body|4 burner gas range|Static oven|Thermostatic control',
     '', '', '', 'active',
     '11-1', '11-2', '11-3', '11-4', '', ''],
]

for row_num, sample in enumerate(samples, 3):
    for col, val in enumerate(sample, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.border = thin_border
        cell.alignment = wrap

ws.freeze_panes = 'A3'
ws.auto_filter.ref = f'A1:{get_column_letter(len(columns))}1'
ws.row_dimensions[1].height = 35
ws.row_dimensions[2].height = 30

# ===== SAYFA 2: ACIKLAMA =====
ws2 = wb.create_sheet('Aciklama')
ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 90

info = [
    ('GASTROTECH URUN YUKLEME SABLONU', '', True, 14),
    ('', '', False, 11),
    ('ZORUNLU ALANLAR', '', True, 12),
    ('Kategori', 'Mevcut kategoriler: Pisirme Ekipmanlari, Firinlar, Sogutma Uniteleri, Buz Makineleri, Hazirlik Ekipmanlari, Kafeterya Ekipmanlari, Tamamlayici Ekipmanlar, Bulashane, Camasirhane, Aksesuarlar', False, 11),
    ('Seri', 'Kategoriye ait seri adi. Ornekler: 600 Serisi, 700 Serisi, 900 Serisi, Drop-in Serisi, PRIME, NEVO, MAESTRO Serisi, MIX, GR, Kitchen Aid, Scotsman, Basic Serisi, Premium Serisi, B Serisi', False, 11),
    ('Urun Adi TR', 'Turkce urun adi. Ayni urunun farkli varyantlari ayni ada sahip olabilir.', False, 11),
    ('Model Kodu', 'Her varyant icin BENZERSIZ model kodu. Ornek: GKO6010, EKF7020, PRIME061EPM', False, 11),
    ('Foto Referansi', 'PDF katalog sayfa numarasi ve o sayfadaki gorsel sirasi. Format: SAYFA-SIRA (ornek: 9-1 = sayfa 9, 1. gorsel)', False, 11),
    ('', '', False, 11),
    ('TEKNIK SPEC ALANLARI', '', True, 12),
    ('Boyutlar', 'GxDxY formatinda, mm cinsinden. Ornek: 800x700x850', False, 11),
    ('Agirlik', 'kg cinsinden net agirlik', False, 11),
    ('Guc', 'Watt veya kW cinsinden. Ornek: 12 (kW) veya 5500 (W)', False, 11),
    ('Voltaj', 'Tam voltaj bilgisi. Ornek: 220-230V 1N AC 50Hz veya 380-415V 3N AC 50Hz', False, 11),
    ('Kapasite', 'Urun kapasitesi. Ornek: 40L, 20xGN1/1, 150 kisi/saat', False, 11),
    ('Yakit Tipi', 'Elektrik, Dogalgaz, LPG veya kombinasyon', False, 11),
    ('Ek Spec Alanlar', '3 adet anahtar-deger cifti. Ornek: pilot_flame=evet, thermostat=50-300C, oven_type=konveksiyonel', False, 11),
    ('', '', False, 11),
    ('OZELLIKLER VE ACIKLAMA', '', True, 12),
    ('Genel Ozellikler', 'Birden fazla ozellik | (pipe) karakteri ile ayrilir. Her | bir madde isaretine donusur.', False, 11),
    ('Uzun Aciklama', 'Serbest metin. Urunun detayli tanitimi.', False, 11),
    ('', '', False, 11),
    ('FOTO REFERANS SISTEMI', '', True, 12),
    ('', 'PDF katalogdaki her gorsel icin SAYFA-SIRA formati kullanilir.', False, 11),
    ('', 'SAYFA = PDF sayfa numarasi (1, 2, 3...)', False, 11),
    ('', 'SIRA = O sayfadaki gorsel sirasi, soldan saga ve yukardan asagi (1, 2, 3...)', False, 11),
    ('', 'Ornek: 9-1 = Sayfa 9 deki 1. gorsel (sol ust)', False, 11),
    ('', '         9-2 = Sayfa 9 deki 2. gorsel (sag ust veya sol alt)', False, 11),
    ('', 'Bir urun icin 5 foto referansi girilebilir.', False, 11),
    ('', '', False, 11),
    ('VARYANT SISTEMI', '', True, 12),
    ('', 'Ayni urunun farkli boyut/tip/yakit varyantlari icin:', False, 11),
    ('', '- Her varyant AYRI SATIR olarak girilir', False, 11),
    ('', '- Urun Adi TR ve Seri AYNI kalir', False, 11),
    ('', '- Model Kodu FARKLI olur', False, 11),
    ('', '- Sistem ayni adi + seriyi gruplayarak tek urun altinda birlestirir', False, 11),
    ('', '', False, 11),
    ('DURUM DEGERLERI', '', True, 12),
    ('active', 'Urun sitede yayinda', False, 11),
    ('draft', 'Taslak - sitede gorunmez', False, 11),
    ('archived', 'Arsivlenmis - sitede gorunmez', False, 11),
]

for row, (key, val, bold, size) in enumerate(info, 1):
    cell_a = ws2.cell(row=row, column=1, value=key)
    cell_a.font = Font(bold=bold, size=size)
    cell_b = ws2.cell(row=row, column=2, value=val)
    cell_b.font = Font(size=11)
    cell_b.alignment = wrap

output = r'C:\gastrotech.com.tr.0101\gastrotech.com_cursor\backend\urun_yukleme_sablonu.xlsx'
wb.save(output)
print(f'Sablon olusturuldu: {output}')
