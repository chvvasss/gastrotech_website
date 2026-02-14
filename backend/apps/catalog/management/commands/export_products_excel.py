"""
Export all products to Excel file with name, code, series, and category.

Usage:
    python manage.py export_products_excel
    python manage.py export_products_excel --output custom_filename.xlsx
"""

from django.core.management.base import BaseCommand

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from apps.catalog.models import Product, Variant


class Command(BaseCommand):
    help = "Export all products to Excel file with basic information"

    def add_arguments(self, parser):
        parser.add_argument(
            "--output",
            "-o",
            type=str,
            default="urun_listesi.xlsx",
            help="Output filename (default: urun_listesi.xlsx)",
        )

    def handle(self, *args, **options):
        if not OPENPYXL_AVAILABLE:
            self.stderr.write(
                self.style.ERROR(
                    "openpyxl is not installed. Please install it with: pip install openpyxl"
                )
            )
            return

        output_file = options["output"]

        self.stdout.write("Ürünler Excel'e aktarılıyor...")

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ürün Listesi"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers
        headers = [
            "Sıra No",
            "Ürün Adı (Türkçe)",
            "Ürün Adı (İngilizce)",
            "Model Kodu",
            "Seri",
            "Kategori",
            "Alt Kategori",
            "Durum",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Fetch all products with related data
        products = Product.objects.select_related(
            "series",
            "series__category",
            "series__category__parent",
            "category",
            "category__parent",
        ).prefetch_related("variants").order_by(
            "series__category__name",
            "series__name",
            "title_tr",
        )

        row_num = 2
        product_count = 0
        variant_count = 0

        for product in products:
            # Get category info
            category = product.category or (product.series.category if product.series else None)

            if category:
                if category.parent:
                    main_category = category.parent.name
                    sub_category = category.name
                else:
                    main_category = category.name
                    sub_category = ""
            else:
                main_category = ""
                sub_category = ""

            # Get series name
            series_name = product.series.name if product.series else ""

            # Get status display
            status_display = {
                "draft": "Taslak",
                "active": "Aktif",
                "archived": "Arşivlenmiş",
            }.get(product.status, product.status)

            # Get variants (model codes)
            variants = product.variants.all()

            if variants.exists():
                for variant in variants:
                    ws.cell(row=row_num, column=1, value=row_num - 1).border = thin_border
                    ws.cell(row=row_num, column=2, value=product.title_tr or product.name).border = thin_border
                    ws.cell(row=row_num, column=3, value=product.title_en or "").border = thin_border
                    ws.cell(row=row_num, column=4, value=variant.model_code).border = thin_border
                    ws.cell(row=row_num, column=5, value=series_name).border = thin_border
                    ws.cell(row=row_num, column=6, value=main_category).border = thin_border
                    ws.cell(row=row_num, column=7, value=sub_category).border = thin_border
                    ws.cell(row=row_num, column=8, value=status_display).border = thin_border
                    row_num += 1
                    variant_count += 1
            else:
                # Product without variants
                ws.cell(row=row_num, column=1, value=row_num - 1).border = thin_border
                ws.cell(row=row_num, column=2, value=product.title_tr or product.name).border = thin_border
                ws.cell(row=row_num, column=3, value=product.title_en or "").border = thin_border
                ws.cell(row=row_num, column=4, value="-").border = thin_border
                ws.cell(row=row_num, column=5, value=series_name).border = thin_border
                ws.cell(row=row_num, column=6, value=main_category).border = thin_border
                ws.cell(row=row_num, column=7, value=sub_category).border = thin_border
                ws.cell(row=row_num, column=8, value=status_display).border = thin_border
                row_num += 1

            product_count += 1

        # Auto-adjust column widths
        column_widths = {
            1: 10,   # Sıra No
            2: 50,   # Ürün Adı TR
            3: 50,   # Ürün Adı EN
            4: 20,   # Model Kodu
            5: 30,   # Seri
            6: 30,   # Kategori
            7: 30,   # Alt Kategori
            8: 15,   # Durum
        }

        for col, width in column_widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save workbook
        wb.save(output_file)

        self.stdout.write(
            self.style.SUCCESS(
                f"\nExport tamamlandı!\n"
                f"  - Toplam ürün: {product_count}\n"
                f"  - Toplam satır (varyantlar dahil): {row_num - 2}\n"
                f"  - Dosya: {output_file}"
            )
        )
