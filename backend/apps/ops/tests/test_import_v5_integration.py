"""
Integration tests for V5 import system behavior.

These tests verify the actual import behavior end-to-end:
- Variant name defaults
- Smart mode candidate creation
- DB verification structure
- Series-category mismatch validation
- Status field defaults
- Required field validation

All tests use in-memory Excel files and actual database operations.
"""

import io
import openpyxl
from decimal import Decimal
from django.test import TestCase

from apps.catalog.models import Category, Series, Brand, Product, Variant
from apps.ops.services.unified_import import UnifiedImportService
from apps.ops.models import ImportJob


class TestImportV5Integration(TestCase):
    """Integration tests for V5 import system behavior."""

    def setUp(self):
        """Clean database before each test."""
        Variant.objects.all().delete()
        Product.objects.all().delete()
        Series.objects.all().delete()
        Category.objects.all().delete()
        Brand.objects.all().delete()
        ImportJob.objects.all().delete()

    def _create_excel_bytes(self, products_data=None, variants_data=None):
        """
        Create in-memory Excel file with Products and Variants sheets.

        Args:
            products_data: List of dicts with product row data
            variants_data: List of dicts with variant row data

        Returns:
            bytes: Excel file content
        """
        wb = openpyxl.Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create Products sheet
        if products_data is not None:
            ws_products = wb.create_sheet("Products")
            # V5 header row
            products_headers = [
                "Brand", "Category", "Series", "Product Name", "Product Slug",
                "Title TR", "Title EN", "Status", "Is Featured",
                "Long Description", "General Features", "Short Specs", "Taxonomy"
            ]
            ws_products.append(products_headers)

            for row_data in products_data:
                ws_products.append([
                    row_data.get('Brand', ''),
                    row_data.get('Category', ''),
                    row_data.get('Series', ''),
                    row_data.get('Product Name', ''),
                    row_data.get('Product Slug', ''),
                    row_data.get('Title TR', ''),
                    row_data.get('Title EN', ''),
                    row_data.get('Status', ''),
                    row_data.get('Is Featured', ''),
                    row_data.get('Long Description', ''),
                    row_data.get('General Features', ''),
                    row_data.get('Short Specs', ''),
                    row_data.get('Taxonomy', ''),
                ])

        # Create Variants sheet
        if variants_data is not None:
            ws_variants = wb.create_sheet("Variants")
            # V5 header row (with Spec: columns if needed)
            variants_headers = [
                "Product Slug", "Model Code", "Variant Name TR", "Variant Name EN",
                "SKU", "Dimensions", "Weight", "List Price", "Stock Qty"
            ]

            # Check if any variant has specs
            all_spec_keys = set()
            for row_data in variants_data:
                if 'specs' in row_data:
                    all_spec_keys.update(row_data['specs'].keys())

            for spec_key in sorted(all_spec_keys):
                variants_headers.append(f"Spec:{spec_key}")

            ws_variants.append(variants_headers)

            for row_data in variants_data:
                row = [
                    row_data.get('Product Slug', ''),
                    row_data.get('Model Code', ''),
                    row_data.get('Variant Name TR', ''),
                    row_data.get('Variant Name EN', ''),
                    row_data.get('SKU', ''),
                    row_data.get('Dimensions', ''),
                    row_data.get('Weight', ''),
                    row_data.get('List Price', ''),
                    row_data.get('Stock Qty', ''),
                ]

                # Add spec values
                if 'specs' in row_data:
                    for spec_key in sorted(all_spec_keys):
                        row.append(row_data['specs'].get(spec_key, ''))

                ws_variants.append(row)

        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.read()

    def test_variant_name_defaults_to_product_title_when_blank(self):
        """
        Test that blank Variant Name TR/EN default to Product Title TR/EN.

        V5 Contract: Variant names are NOT required - they default to Product titles.
        """
        # Create product data with specific titles
        products_data = [{
            'Brand': 'acme',
            'Category': 'electronics',
            'Series': 'premium',
            'Product Name': 'Test Product',
            'Product Slug': 'test-product',
            'Title TR': 'Test Product TR',
            'Title EN': 'Test Product EN',
        }]

        # Create variant with BLANK names
        variants_data = [{
            'Product Slug': 'test-product',
            'Model Code': 'TP-001',
            'Variant Name TR': '',  # Blank - should default
            'Variant Name EN': '',  # Blank - should default
            'SKU': 'SKU001',
        }]

        excel_bytes = self._create_excel_bytes(products_data, variants_data)

        # Validate with SMART mode (to create missing entities)
        service = UnifiedImportService(mode='smart')
        report = service.validate(excel_bytes, 'test.xlsx')

        # Should pass validation
        assert report['status'] in ['validation_passed', 'validation_warnings']

        # Check that variants_data has correct defaults
        assert len(report['variants_data']) == 1
        variant_data = report['variants_data'][0]
        assert variant_data['name_tr'] == 'Test Product TR', \
            f"Expected name_tr='Test Product TR', got '{variant_data['name_tr']}'"
        assert variant_data['name_en'] == 'Test Product EN', \
            f"Expected name_en='Test Product EN', got '{variant_data['name_en']}'"

    def test_category_series_brand_creation_smart_mode(self):
        """
        Test that smart mode creates missing entities and db_verify confirms them.

        V5 Contract: Smart mode creates candidates, commit verifies DB writes.
        """
        # Start with empty DB (setup fixture already cleaned it)
        assert Category.objects.count() == 0
        assert Series.objects.count() == 0
        assert Brand.objects.count() == 0

        # Import with missing entities
        products_data = [{
            'Brand': 'acme',
            'Category': 'electronics',
            'Series': 'premium',
            'Product Name': 'Test Product',
            'Product Slug': 'test-product',
            'Title TR': 'Test Title TR',
        }]

        variants_data = [{
            'Product Slug': 'test-product',
            'Model Code': 'TP-001',
        }]

        excel_bytes = self._create_excel_bytes(products_data, variants_data)

        # Phase 1: Validate with SMART mode
        service = UnifiedImportService(mode='smart')
        report = service.validate(excel_bytes, 'test.xlsx')

        # Should have candidates
        assert len(report['candidates']['categories']) > 0
        assert len(report['candidates']['brands']) > 0
        assert len(report['candidates']['series']) > 0

        # Find category candidate
        category_candidates = [c for c in report['candidates']['categories'] if c['slug'] == 'electronics']
        assert len(category_candidates) == 1

        # Find brand candidate
        brand_candidates = [b for b in report['candidates']['brands'] if b['slug'] == 'acme']
        assert len(brand_candidates) == 1

        # Find series candidate
        series_candidates = [s for s in report['candidates']['series'] if s['slug'] == 'premium']
        assert len(series_candidates) == 1

        # Phase 2: Commit (create ImportJob first)
        from apps.catalog.models import Media

        input_media = Media.objects.create(
            kind='file',
            filename='test.xlsx',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            bytes=excel_bytes,
            size_bytes=len(excel_bytes),
        )

        snapshot_data = report.get('snapshot')
        snapshot_media = Media.objects.get(id=snapshot_data['media_id'])

        job = ImportJob.objects.create(
            kind='catalog_import',
            mode='smart',
            status='pending',
            input_file=input_media,
            snapshot_file=snapshot_media,
            snapshot_hash=snapshot_data['hash'],
        )

        # Store report_json for commit validation check
        job.report_json = report
        job.save()

        # Commit
        commit_result = service.commit(str(job.id), allow_partial=False)

        # Verify commit result structure
        assert commit_result['status'] == 'success'
        assert 'db_verify' in commit_result
        assert commit_result['db_verify']['enabled'] is True
        assert commit_result['db_verify']['created_entities_found_in_db'] is True

        # Verify entities exist in DB
        assert Category.objects.filter(slug='electronics').exists()
        assert Brand.objects.filter(slug='acme').exists()
        assert Series.objects.filter(slug='premium').exists()
        assert Product.objects.filter(slug='test-product').exists()
        assert Variant.objects.filter(model_code='TP-001').exists()

    def test_commit_db_verify_structure(self):
        """
        Test that commit returns proper db_verify structure.

        V5 Contract: db_verify must include all verification fields.
        """
        # Create minimal valid import
        products_data = [{
            'Brand': 'test-brand',
            'Category': 'test-category',
            'Series': 'test-series',
            'Product Name': 'Product',
            'Product Slug': 'product-slug',
            'Title TR': 'Product Title',
        }]

        variants_data = [{
            'Product Slug': 'product-slug',
            'Model Code': 'P-001',
        }]

        excel_bytes = self._create_excel_bytes(products_data, variants_data)

        # Validate
        service = UnifiedImportService(mode='smart')
        report = service.validate(excel_bytes, 'test.xlsx')

        assert report['status'] in ['validation_passed', 'validation_warnings']

        # Create ImportJob
        from apps.catalog.models import Media

        input_media = Media.objects.create(
            kind='file',
            filename='test.xlsx',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            bytes=excel_bytes,
            size_bytes=len(excel_bytes),
        )

        snapshot_data = report.get('snapshot')
        snapshot_media = Media.objects.get(id=snapshot_data['media_id'])

        job = ImportJob.objects.create(
            kind='catalog_import',
            mode='smart',
            status='pending',
            input_file=input_media,
            snapshot_file=snapshot_media,
            snapshot_hash=snapshot_data['hash'],
        )
        job.report_json = report
        job.save()

        # Commit
        commit_result = service.commit(str(job.id))

        # Assert db_verify structure
        assert 'db_verify' in commit_result
        db_verify = commit_result['db_verify']

        assert 'enabled' in db_verify
        assert db_verify['enabled'] is True

        assert 'created_entities_found_in_db' in db_verify
        assert db_verify['created_entities_found_in_db'] is True

        # Check for slug/code lists
        assert 'created_category_slugs' in db_verify
        assert 'created_brand_slugs' in db_verify
        assert 'created_series_slugs' in db_verify
        assert 'created_product_slugs' in db_verify
        assert 'created_variant_model_codes' in db_verify

    def test_series_category_mismatch_strict_mode(self):
        """
        Test that series-category mismatch produces error in strict mode.

        V5 Contract: If Series exists and belongs to Category A,
        but file specifies Category B, validation must fail.
        """
        # Create existing entities
        electronics_cat = Category.objects.create(slug='electronics', name='Electronics')
        furniture_cat = Category.objects.create(slug='furniture', name='Furniture')
        premium_series = Series.objects.create(
            slug='premium',
            name='Premium Series',
            category=electronics_cat  # Belongs to electronics
        )

        # Import with mismatch: Series=premium (electronics) but Category=furniture
        products_data = [{
            'Brand': 'acme',
            'Category': 'furniture',  # MISMATCH!
            'Series': 'premium',       # Belongs to electronics
            'Product Name': 'Test Product',
            'Product Slug': 'test-product',
            'Title TR': 'Test Title',
        }]

        excel_bytes = self._create_excel_bytes(products_data, None)

        # Validate with STRICT mode
        service = UnifiedImportService(mode='strict')
        report = service.validate(excel_bytes, 'test.xlsx')

        # Should fail validation
        assert report['status'] == 'failed_validation'

        # Check for specific error
        mismatch_errors = [
            issue for issue in report['issues']
            if issue.get('code') == 'series_category_mismatch'
        ]
        assert len(mismatch_errors) > 0, "Expected series_category_mismatch error"

        error = mismatch_errors[0]
        assert 'premium' in error['message'].lower()
        assert 'electronics' in error['message'].lower()
        assert 'furniture' in error['message'].lower()

    def test_status_defaults_to_active(self):
        """
        Test that blank Status field defaults to 'active'.

        V5 Contract: Status defaults to 'active', not 'draft'.
        """
        products_data = [{
            'Brand': 'acme',
            'Category': 'electronics',
            'Series': 'premium',
            'Product Name': 'Test Product',
            'Product Slug': 'test-product',
            'Title TR': 'Test Title',
            'Status': '',  # Blank - should default to 'active'
        }]

        excel_bytes = self._create_excel_bytes(products_data, None)

        # Validate
        service = UnifiedImportService(mode='smart')
        report = service.validate(excel_bytes, 'test.xlsx')

        assert report['status'] in ['validation_passed', 'validation_warnings']

        # Check products_data
        assert len(report['products_data']) == 1
        product_data = report['products_data'][0]
        assert product_data['status'] == 'active', \
            f"Expected status='active' for blank Status field, got '{product_data['status']}'"

    def test_required_fields_brand_category(self):
        """
        Test that Category is required but Brand is optional.

        V5.1 Contract: Category and Series are REQUIRED. Brand is OPTIONAL.
        """
        # Test missing Brand - should PASS (Brand is optional now)
        Category.objects.create(slug='electronics', name='Electronics')
        Series.objects.create(
            slug='premium',
            name='Premium',
            category=Category.objects.get(slug='electronics')
        )

        products_data_no_brand = [{
            'Brand': '',  # Missing - this is OK now!
            'Category': 'electronics',
            'Series': 'premium',
            'Product Name': 'Test Product',
            'Product Slug': 'test-product-1',
            'Title TR': 'Test Title',
        }]

        excel_bytes = self._create_excel_bytes(products_data_no_brand, None)

        service = UnifiedImportService(mode='strict')
        report = service.validate(excel_bytes, 'test.xlsx')

        # Should PASS validation (Brand is optional)
        assert report['status'] in ['validation_passed', 'ready_to_commit'], f"Expected validation_passed, got {report['status']}"

        # No Brand errors should exist
        brand_errors = [
            issue for issue in report['issues']
            if 'brand' in issue.get('message', '').lower() and issue.get('severity') == 'error'
        ]
        assert len(brand_errors) == 0, "Brand should be optional, no errors expected"

        # Test missing Category
        products_data_no_category = [{
            'Brand': 'acme',
            'Category': '',  # Missing
            'Series': 'premium',
            'Product Name': 'Test Product',
            'Product Slug': 'test-product-2',
            'Title TR': 'Test Title',
        }]

        excel_bytes = self._create_excel_bytes(products_data_no_category, None)

        service = UnifiedImportService(mode='strict')
        report = service.validate(excel_bytes, 'test.xlsx')

        # Should fail validation
        assert report['status'] == 'failed_validation'

        # Check for Category error
        category_errors = [
            issue for issue in report['issues']
            if 'category' in issue.get('message', '').lower() and issue.get('severity') == 'error'
        ]
        assert len(category_errors) > 0, "Expected 'Category is required' error"
        assert 'required' in category_errors[0]['message'].lower()

    def test_brand_optional_allows_product_import(self):
        """
        Test that products can be imported without a brand (brand=NULL).

        V5.1 Feature: Brand field is optional. Product should be created with brand=NULL.
        """
        # Create necessary objects
        category = Category.objects.create(slug='electronics', name='Electronics')
        series = Series.objects.create(slug='premium', name='Premium', category=category)

        # Product data WITHOUT brand
        products_data = [{
            'Brand': '',  # Empty/blank brand
            'Category': 'electronics',
            'Series': 'premium',
            'Product Name': 'Generic Product',
            'Product Slug': 'generic-product',
            'Title TR': 'Genel Ürün',
        }]

        variants_data = [{
            'Product Slug': 'generic-product',
            'Model Code': 'GP-001',
        }]

        excel_bytes = self._create_excel_bytes(products_data, variants_data)

        # Validate
        service = UnifiedImportService(mode='strict')
        report = service.validate(excel_bytes, 'test.xlsx')

        # Validation should pass
        assert report['status'] in ['validation_passed', 'ready_to_commit'], f"Validation should pass, got: {report['status']}"

        # Should have no validation errors
        assert len([i for i in report.get('issues', []) if i.get('severity') == 'error']) == 0

        # Create ImportJob for commit
        from apps.catalog.models import Media

        input_media = Media.objects.create(
            kind='file',
            filename='test.xlsx',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            bytes=excel_bytes,
            size_bytes=len(excel_bytes),
        )

        snapshot_data = report.get('snapshot')
        snapshot_media = Media.objects.get(id=snapshot_data['media_id'])

        job = ImportJob.objects.create(
            kind='catalog_import',
            mode='strict',
            status='pending',
            input_file=input_media,
            snapshot_file=snapshot_media,
            snapshot_hash=snapshot_data['hash'],
        )

        job.report_json = report
        job.save()

        # Commit
        commit_report = service.commit(str(job.id))

        assert commit_report['status'] == 'success', f"Commit should succeed, got: {commit_report.get('status')}"

        # Verify product in database has brand=NULL
        product = Product.objects.get(slug='generic-product')
        assert product.brand is None, "Product should have no brand (NULL)"
        assert product.series == series
        assert product.name == 'Generic Product'

    def test_brand_candidate_only_when_brand_provided(self):
        """
        Test that brand candidates are only created when brand is provided but not found.

        Behavior:
        - Brand blank: No candidate, no error
        - Brand provided but missing in STRICT: Error
        - Brand provided but missing in SMART: Candidate created
        """
        category = Category.objects.create(slug='electronics', name='Electronics')
        series = Series.objects.create(slug='premium', name='Premium', category=category)

        # Test 1: Brand blank - no candidate
        products_data_blank = [{
            'Brand': '',
            'Category': 'electronics',
            'Series': 'premium',
            'Product Name': 'Product A',
            'Product Slug': 'product-a',
            'Title TR': 'Ürün A',
        }]

        excel_bytes = self._create_excel_bytes(products_data_blank, None)
        service = UnifiedImportService(mode='smart')
        report = service.validate(excel_bytes, 'test.xlsx')

        assert report['status'] in ['validation_passed', 'ready_to_commit']
        assert len(report['candidates']['brands']) == 0, "No brand candidate when brand is blank"

        # Test 2: Brand provided but not found in STRICT - error
        products_data_missing_strict = [{
            'Brand': 'nonexistent-brand',
            'Category': 'electronics',
            'Series': 'premium',
            'Product Name': 'Product B',
            'Product Slug': 'product-b',
            'Title TR': 'Ürün B',
        }]

        excel_bytes = self._create_excel_bytes(products_data_missing_strict, None)
        service = UnifiedImportService(mode='strict')
        report = service.validate(excel_bytes, 'test.xlsx')

        assert report['status'] == 'failed_validation'
        brand_errors = [
            issue for issue in report['issues']
            if 'nonexistent-brand' in str(issue.get('value', ''))
        ]
        assert len(brand_errors) > 0, "Should error for missing brand in STRICT mode"

        # Test 3: Brand provided but not found in SMART - candidate
        excel_bytes = self._create_excel_bytes(products_data_missing_strict, None)
        service = UnifiedImportService(mode='smart')
        report = service.validate(excel_bytes, 'test.xlsx')

        assert report['status'] in ['validation_passed', 'ready_to_commit']
        assert len(report['candidates']['brands']) == 1, "Should create brand candidate in SMART mode"
        assert report['candidates']['brands'][0]['slug'] == 'nonexistent-brand'

    def test_series_category_exact_match_passes(self):
        """
        Test that series-category exact match passes validation.

        V5.1 Feature: series_category_match_type should be 'exact' when slugs match.
        """
        # Create matching category and series
        category = Category.objects.create(slug='electronics', name='Electronics')
        series = Series.objects.create(slug='premium', name='Premium', category=category)

        products_data = [{
            'Brand': '',
            'Category': 'electronics',  # EXACT match with series category
            'Series': 'premium',
            'Product Name': 'Test Product',
            'Product Slug': 'test-product',
            'Title TR': 'Test Title',
        }]

        excel_bytes = self._create_excel_bytes(products_data, None)
        service = UnifiedImportService(mode='strict')
        report = service.validate(excel_bytes, 'test.xlsx')

        # Should pass validation
        assert report['status'] in ['validation_passed', 'ready_to_commit'], \
            f"Expected validation_passed, got {report['status']}"

        # Should have match_type='exact'
        assert len(report['products_data']) == 1
        assert report['products_data'][0].get('series_category_match_type') == 'exact', \
            f"Expected match_type='exact', got '{report['products_data'][0].get('series_category_match_type')}'"

    def test_series_category_ancestor_match_passes(self):
        """
        Test that series in parent category allows product in child category.

        V5.1 Feature: Products can be in subcategories of the series' category.
        series_category_match_type should be 'ancestor'.
        """
        # Create parent category
        parent_cat = Category.objects.create(slug='firinlar', name='Fırınlar')
        # Create child category under parent
        child_cat = Category.objects.create(
            slug='kombi-firinlar',
            name='Kombi Fırınlar',
            parent=parent_cat
        )
        # Create series in PARENT category
        series = Series.objects.create(slug='i-combi', name='iCombi', category=parent_cat)

        products_data = [{
            'Brand': '',
            'Category': 'kombi-firinlar',  # CHILD category (descendant of series category)
            'Series': 'i-combi',           # Series is in PARENT category (firinlar)
            'Product Name': 'iCombi Pro',
            'Product Slug': 'icombi-pro',
            'Title TR': 'iCombi Pro',
        }]

        excel_bytes = self._create_excel_bytes(products_data, None)
        service = UnifiedImportService(mode='strict')
        report = service.validate(excel_bytes, 'test.xlsx')

        # Should PASS - series category is ancestor of product category
        assert report['status'] in ['validation_passed', 'ready_to_commit'], \
            f"Expected validation_passed, got {report['status']}. Issues: {report.get('issues', [])}"

        # Should have NO series_category_mismatch errors
        mismatch_errors = [i for i in report['issues'] if i.get('code') == 'series_category_mismatch']
        assert len(mismatch_errors) == 0, \
            f"Expected no mismatch errors, got {mismatch_errors}"

        # Should have match_type='ancestor'
        assert len(report['products_data']) == 1
        assert report['products_data'][0].get('series_category_match_type') == 'ancestor', \
            f"Expected match_type='ancestor', got '{report['products_data'][0].get('series_category_match_type')}'"

    def test_series_category_mismatch_unrelated_fails(self):
        """
        Test that series in unrelated category fails validation.

        V5.1: When categories are NOT in an ancestor-descendant relationship,
        validation must fail with series_category_mismatch error.
        """
        # Create two UNRELATED categories (neither is ancestor of the other)
        cat_electronics = Category.objects.create(slug='electronics', name='Electronics')
        cat_furniture = Category.objects.create(slug='furniture', name='Furniture')

        # Series belongs to electronics
        series = Series.objects.create(
            slug='premium',
            name='Premium',
            category=cat_electronics
        )

        products_data = [{
            'Brand': '',
            'Category': 'furniture',  # UNRELATED to series' category (electronics)
            'Series': 'premium',      # belongs to electronics
            'Product Name': 'Test Product',
            'Product Slug': 'test-product',
            'Title TR': 'Test Title',
        }]

        excel_bytes = self._create_excel_bytes(products_data, None)
        service = UnifiedImportService(mode='strict')
        report = service.validate(excel_bytes, 'test.xlsx')

        # Should FAIL - categories are unrelated
        assert report['status'] == 'failed_validation', \
            f"Expected failed_validation for unrelated categories, got {report['status']}"

        # Should have series_category_mismatch error
        mismatch_errors = [i for i in report['issues'] if i.get('code') == 'series_category_mismatch']
        assert len(mismatch_errors) > 0, "Expected series_category_mismatch error"

        # Check error message mentions both categories
        error_msg = mismatch_errors[0]['message'].lower()
        assert 'electronics' in error_msg, "Error should mention series category"
        assert 'furniture' in error_msg, "Error should mention product category"
