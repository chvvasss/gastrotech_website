"""
Tests for V5 Template Contract Enforcement.

These tests ensure the import system stays aligned with bulk_upload_template (5).xlsx.
ANY change to template structure MUST update TEMPLATE_CONTRACT.md and these tests.
"""

import openpyxl
import pytest
from pathlib import Path


@pytest.mark.django_db
class TestTemplateContractV5:
    """
    Critical tests to prevent template contract violations.

    If these tests fail, the template and backend are out of sync - BLOCKING issue.
    """

    @pytest.fixture
    def template_path(self):
        """Get path to canonical V5 template."""
        base_path = Path(__file__).parent.parent.parent.parent.parent
        template_path = base_path / "SABLON" / "bulk_upload_template (5).xlsx"
        assert template_path.exists(), f"V5 template not found at {template_path}"
        return template_path

    def test_template_file_exists(self, template_path):
        """V5 template file must exist at expected location."""
        assert template_path.exists()
        assert template_path.suffix == ".xlsx"

    def test_products_sheet_columns(self, template_path):
        """
        Products sheet MUST have EXACT V5 columns in EXACT order.

        Contract: TEMPLATE_CONTRACT.md Section "Products Sheet"
        """
        wb = openpyxl.load_workbook(template_path, data_only=True)
        assert "Products" in wb.sheetnames, "Products sheet missing"

        ws = wb["Products"]
        headers = [cell.value for cell in ws[1] if cell.value]

        expected_columns = [
            "Brand",
            "Category",
            "Series",
            "Product Name",
            "Product Slug",
            "Title TR",
            "Title EN",
            "Status",
            "Is Featured",
            "Long Description",
            "General Features",
            "Short Specs",
            "Taxonomy",
        ]

        assert headers == expected_columns, (
            f"Products sheet columns mismatch!\n"
            f"Expected: {expected_columns}\n"
            f"Got: {headers}\n"
            f"Contract violation: Update TEMPLATE_CONTRACT.md if this is intentional"
        )

    def test_variants_sheet_columns(self, template_path):
        """
        Variants sheet MUST have V5 required columns (exact order for first 9, Spec:* dynamic).

        Contract: TEMPLATE_CONTRACT.md Section "Variants Sheet"
        """
        wb = openpyxl.load_workbook(template_path, data_only=True)
        assert "Variants" in wb.sheetnames, "Variants sheet missing"

        ws = wb["Variants"]
        headers = [cell.value for cell in ws[1] if cell.value]

        # V5 required columns in exact order
        required_columns = [
            "Product Slug",
            "Model Code",
            "Variant Name TR",
            "Variant Name EN",
            "SKU",
            "Dimensions",
            "Weight",
            "List Price",
            "Stock Qty",
        ]

        # Check first 9 columns match exactly
        assert headers[:9] == required_columns, (
            f"Variants sheet required columns mismatch!\n"
            f"Expected: {required_columns}\n"
            f"Got: {headers[:9]}\n"
            f"Contract violation: Update TEMPLATE_CONTRACT.md if this is intentional"
        )

        # Check remaining columns are Spec:* format
        spec_columns = headers[9:]
        for col in spec_columns:
            assert col.startswith("Spec:"), (
                f"Column '{col}' after required columns must start with 'Spec:'\n"
                f"Got: {spec_columns}\n"
                f"Contract: All dynamic columns must be Spec:*"
            )

    def test_sheet_count_and_names(self, template_path):
        """
        Template MUST have exactly 4 sheets: Products, Variants, Reference Data, Instructions.

        Contract: TEMPLATE_CONTRACT.md Section "Sheet Structure"
        """
        wb = openpyxl.load_workbook(template_path, data_only=True)
        expected_sheets = ["Products", "Variants", "Reference Data", "Instructions"]

        assert wb.sheetnames == expected_sheets, (
            f"Template sheets mismatch!\n"
            f"Expected: {expected_sheets}\n"
            f"Got: {wb.sheetnames}\n"
            f"Contract violation: Template must have exactly these 4 sheets"
        )

    def test_reference_data_sheet_columns(self, template_path):
        """
        Reference Data sheet MUST have exactly 5 columns.

        Contract: TEMPLATE_CONTRACT.md Section "Reference Data Sheet"
        """
        wb = openpyxl.load_workbook(template_path, data_only=True)
        assert "Reference Data" in wb.sheetnames

        ws = wb["Reference Data"]
        headers = [cell.value for cell in ws[1] if cell.value]

        expected_columns = [
            "Valid Brands",
            "Valid Categories",
            "Valid Series",
            "Valid Status",
            "Valid Is Featured",
        ]

        assert headers == expected_columns, (
            f"Reference Data sheet columns mismatch!\n"
            f"Expected: {expected_columns}\n"
            f"Got: {headers}\n"
            f"Contract violation: Update TEMPLATE_CONTRACT.md if this is intentional"
        )

    def test_backend_column_mapping_matches_template(self, template_path):
        """
        Backend PRODUCTS_COLUMN_MAP must include all V5 template columns as PRIMARY aliases.

        Contract: Backend mappings must match template column names exactly.
        """
        from apps.ops.services.unified_import import PRODUCTS_COLUMN_MAP

        wb = openpyxl.load_workbook(template_path, data_only=True)
        ws = wb["Products"]
        template_columns = [cell.value for cell in ws[1] if cell.value]

        # Check each template column has a mapping and is the PRIMARY (first) alias
        for col in template_columns:
            # Find which internal field maps to this column
            mapped_field = None
            for field, aliases in PRODUCTS_COLUMN_MAP.items():
                if col in aliases:
                    mapped_field = field
                    # V5 requirement: Template column must be the FIRST alias (canonical)
                    assert aliases[0] == col, (
                        f"Template column '{col}' maps to field '{field}' but is not the first alias!\n"
                        f"Aliases: {aliases}\n"
                        f"Contract: V5 column names must be PRIMARY (first in alias list)"
                    )
                    break

            assert mapped_field is not None, (
                f"Template column '{col}' has NO backend mapping!\n"
                f"Available mappings: {list(PRODUCTS_COLUMN_MAP.keys())}\n"
                f"Contract violation: All template columns must be mapped"
            )

    def test_backend_required_fields_match_template(self):
        """
        Backend PRODUCTS_REQUIRED must match V5 template contract.

        Contract: Brand, Category, Series, Product Name, Product Slug, Title TR are REQUIRED.
        """
        from apps.ops.services.unified_import import PRODUCTS_REQUIRED

        expected_required = [
            'brand_slug',     # Brand
            'category_slug',  # Category
            'series_slug',    # Series
            'name',           # Product Name
            'slug',           # Product Slug
            'title_tr',       # Title TR
        ]

        assert set(PRODUCTS_REQUIRED) == set(expected_required), (
            f"Backend PRODUCTS_REQUIRED mismatch!\n"
            f"Expected: {expected_required}\n"
            f"Got: {PRODUCTS_REQUIRED}\n"
            f"Contract: Template defines Brand/Category/Series/Product Name/Product Slug/Title TR as required"
        )

    def test_backend_variants_required_fields(self):
        """
        Backend VARIANTS_REQUIRED must match V5 contract: ONLY product_slug and model_code.

        Contract: name_tr/name_en are NOT required (default to Product.title_tr/title_en).
        """
        from apps.ops.services.unified_import import VARIANTS_REQUIRED

        expected_required = [
            'product_slug',  # Product Slug
            'model_code',    # Model Code
        ]

        assert VARIANTS_REQUIRED == expected_required, (
            f"Backend VARIANTS_REQUIRED mismatch!\n"
            f"Expected: {expected_required}\n"
            f"Got: {VARIANTS_REQUIRED}\n"
            f"Contract: Variant Name TR/EN are NOT required, default to Product titles"
        )

    def test_template_has_no_extra_sheets(self, template_path):
        """
        Template must not contain deprecated or extra sheets.

        Contract: Only Products, Variants, Reference Data, Instructions allowed.
        """
        wb = openpyxl.load_workbook(template_path, data_only=True)
        allowed_sheets = {"Products", "Variants", "Reference Data", "Instructions"}
        extra_sheets = set(wb.sheetnames) - allowed_sheets

        assert not extra_sheets, (
            f"Template contains extra/deprecated sheets: {extra_sheets}\n"
            f"Contract: Remove any sheets not in {allowed_sheets}"
        )

    def test_products_sheet_not_empty(self, template_path):
        """
        Products sheet must have at least header row.

        Contract: Sheet must be parseable (not completely empty).
        """
        wb = openpyxl.load_workbook(template_path, data_only=True)
        ws = wb["Products"]

        assert ws.max_row >= 1, "Products sheet is empty (no header row)"
        assert ws.max_column >= 13, "Products sheet has too few columns"

    def test_variants_sheet_not_empty(self, template_path):
        """
        Variants sheet must have at least header row.

        Contract: Sheet must be parseable (not completely empty).
        """
        wb = openpyxl.load_workbook(template_path, data_only=True)
        ws = wb["Variants"]

        assert ws.max_row >= 1, "Variants sheet is empty (no header row)"
        assert ws.max_column >= 9, "Variants sheet has too few columns"
