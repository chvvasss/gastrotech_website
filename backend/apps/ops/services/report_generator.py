"""
XLSX Report Generator for Import Jobs.

Generates multi-sheet Excel reports with:
- Summary sheet (counts, status, warnings)
- Issues sheet (validation errors/warnings)
- Data sheet (normalized rows, re-import ready)
- Candidates sheet (missing entities to create)
- Normalization sheet (merges, disambiguations)
"""

import io
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ImportReportGenerator:
    """
    Generates comprehensive XLSX reports for import jobs.
    """

    # Color scheme
    COLOR_HEADER = "4472C4"  # Blue
    COLOR_SUCCESS = "70AD47"  # Green
    COLOR_WARNING = "FFC000"  # Orange
    COLOR_ERROR = "C00000"    # Red
    COLOR_INFO = "7030A0"     # Purple

    def generate(self, report_json: Dict[str, Any]) -> bytes:
        """
        Generate XLSX report from import job report_json.

        Returns:
            bytes: XLSX file content
        """
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Generate sheets
        self._create_summary_sheet(wb, report_json)
        self._create_issues_sheet(wb, report_json)
        self._create_data_sheet(wb, report_json)
        self._create_candidates_sheet(wb, report_json)
        self._create_normalization_sheet(wb, report_json)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    def _create_summary_sheet(self, wb: Workbook, report: Dict):
        """Create summary sheet with key metrics."""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws['A1'] = "Import Job Summary"
        ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
        ws.merge_cells('A1:B1')

        # Status
        row = 3
        ws[f'A{row}'] = "Status"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = report.get('status', 'unknown')
        status_color = self._get_status_color(report.get('status'))
        ws[f'B{row}'].fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")

        # Counts
        row += 2
        ws[f'A{row}'] = "Counts"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1

        counts = report.get('counts', {})
        count_labels = {
            'total_product_rows': 'Total Product Rows',
            'total_variant_rows': 'Total Variant Rows',
            'valid_product_rows': 'Valid Product Rows',
            'valid_variant_rows': 'Valid Variant Rows',
            'error_rows': 'Error Rows',
            'warning_rows': 'Warning Rows',
            'products_to_create': 'Products to Create',
            'products_to_update': 'Products to Update',
            'variants_to_create': 'Variants to Create',
            'variants_to_update': 'Variants to Update',
        }

        for key, label in count_labels.items():
            ws[f'A{row}'] = label
            ws[f'B{row}'] = counts.get(key, 0)
            row += 1

        # Candidates (if smart mode)
        candidates = report.get('candidates', {})
        total_candidates = sum(len(v) for v in candidates.values())

        if total_candidates > 0:
            row += 1
            ws[f'A{row}'] = "Missing Entities (Smart Mode)"
            ws[f'A{row}'].font = Font(size=14, bold=True)
            row += 1

            for entity_type, items in candidates.items():
                if items:
                    ws[f'A{row}'] = entity_type.title()
                    ws[f'B{row}'] = len(items)
                    row += 1

        # Auto-size columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def _create_issues_sheet(self, wb: Workbook, report: Dict):
        """Create issues sheet with all validation errors/warnings."""
        ws = wb.create_sheet("Issues")

        headers = ['Row', 'Severity', 'Code', 'Column', 'Value', 'Message', 'Expected']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        issues = report.get('issues', [])
        for row_idx, issue in enumerate(issues, start=2):
            ws.cell(row=row_idx, column=1, value=issue.get('row'))
            ws.cell(row=row_idx, column=2, value=issue.get('severity', '').upper())
            ws.cell(row=row_idx, column=3, value=issue.get('code'))
            ws.cell(row=row_idx, column=4, value=issue.get('column'))
            ws.cell(row=row_idx, column=5, value=str(issue.get('value', '')))
            ws.cell(row=row_idx, column=6, value=issue.get('message'))
            ws.cell(row=row_idx, column=7, value=issue.get('expected') or '')

            severity_cell = ws.cell(row=row_idx, column=2)
            if issue.get('severity') == 'error':
                severity_cell.fill = PatternFill(start_color=self.COLOR_ERROR, end_color=self.COLOR_ERROR, fill_type="solid")
                severity_cell.font = Font(color="FFFFFF", bold=True)
            elif issue.get('severity') == 'warning':
                severity_cell.fill = PatternFill(start_color=self.COLOR_WARNING, end_color=self.COLOR_WARNING, fill_type="solid")
                severity_cell.font = Font(bold=True)
            elif issue.get('severity') == 'info':
                severity_cell.fill = PatternFill(start_color=self.COLOR_INFO, end_color=self.COLOR_INFO, fill_type="solid")
                severity_cell.font = Font(color="FFFFFF")

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 60
        ws.column_dimensions['G'].width = 40

        ws.freeze_panes = 'A2'

    def _create_data_sheet(self, wb: Workbook, report: Dict):
        """Create data sheet with normalized rows (re-import ready)."""
        ws = wb.create_sheet("Data")

        valid_rows = report.get('valid_rows', [])

        if not valid_rows:
            ws['A1'] = "No valid rows to display"
            return

        # Extract columns from first row
        first_row_data = valid_rows[0]['data']
        columns = list(first_row_data.keys())

        # Header
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_idx, row_info in enumerate(valid_rows, start=2):
            row_data = row_info['data']
            for col_idx, col_name in enumerate(columns, start=1):
                value = row_data.get(col_name)
                # Handle pandas NaN
                if value is None or (isinstance(value, float) and str(value) == 'nan'):
                    value = ''
                ws.cell(row=row_idx, column=col_idx, value=str(value))

        # Auto-size columns
        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

        # Freeze header row
        ws.freeze_panes = 'A2'

    def _create_candidates_sheet(self, wb: Workbook, report: Dict):
        """Create candidates sheet (missing entities to create)."""
        ws = wb.create_sheet("Candidates")

        candidates = report.get('candidates', {})
        total_candidates = sum(len(v) for v in candidates.values())

        if total_candidates == 0:
            ws['A1'] = "No missing entities (strict mode or all entities exist)"
            return

        # Title
        ws['A1'] = "Missing Entities (Smart Mode)"
        ws['A1'].font = Font(size=14, bold=True)

        row = 3

        for entity_type, items in candidates.items():
            if not items:
                continue

            ws[f'A{row}'] = entity_type.title()
            ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color=self.COLOR_INFO, end_color=self.COLOR_INFO, fill_type="solid")
            row += 1

            # Extract columns from first item
            if items:
                headers = list(items[0].keys())
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=header)
                    cell.font = Font(bold=True)

                row += 1

                # Data (V5: format rows as comma-separated string)
                for item in items:
                    for col_idx, key in enumerate(headers, start=1):
                        value = item.get(key, '')
                        # Format rows list as comma-separated string
                        if key == 'rows' and isinstance(value, list):
                            value = ', '.join(str(r) for r in sorted(value))
                        ws.cell(row=row, column=col_idx, value=str(value))
                    row += 1

            row += 1  # Blank row between entity types

        # Auto-size columns
        for col_idx in range(1, 5):
            ws.column_dimensions[get_column_letter(col_idx)].width = 25

    def _create_normalization_sheet(self, wb: Workbook, report: Dict):
        """Create normalization sheet (merges, disambiguations)."""
        ws = wb.create_sheet("Normalization")

        normalization = report.get('normalization', {})

        # Title
        ws['A1'] = "Normalization Summary"
        ws['A1'].font = Font(size=14, bold=True)

        row = 3

        # Empty value normalizations
        ws[f'A{row}'] = "Empty Values Normalized"
        ws[f'B{row}'] = normalization.get('empty_value_normalizations', 0)
        row += 2

        # Continuation rows
        merged = normalization.get('merged_continuation_rows', [])
        ws[f'A{row}'] = "Merged Continuation Rows"
        ws[f'B{row}'] = len(merged)
        row += 1

        if merged:
            ws[f'A{row}'] = "Primary Row"
            ws[f'B{row}'] = "Continuation Row"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            row += 1

            for merge_info in merged:
                ws[f'A{row}'] = merge_info['primary_row']
                ws[f'B{row}'] = merge_info['continuation_row']
                row += 1

        row += 1

        # Disambiguated model codes
        disambiguated = normalization.get('disambiguated_model_codes', [])
        ws[f'A{row}'] = "Disambiguated Model Codes"
        ws[f'B{row}'] = len(disambiguated)
        row += 1

        if disambiguated:
            ws[f'A{row}'] = "Row"
            ws[f'B{row}'] = "Original"
            ws[f'C{row}'] = "New"
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].font = Font(bold=True)
            row += 1

            for disambig_info in disambiguated:
                ws[f'A{row}'] = disambig_info['row']
                ws[f'B{row}'] = disambig_info['original']
                ws[f'C{row}'] = disambig_info['new']
                row += 1

        # Auto-size columns
        for col_idx in range(1, 4):
            ws.column_dimensions[get_column_letter(col_idx)].width = 25

    def _get_status_color(self, status: str) -> str:
        """Get color for status."""
        status_colors = {
            'validation_passed': self.COLOR_SUCCESS,
            'validation_warnings': self.COLOR_WARNING,
            'failed_validation': self.COLOR_ERROR,
            'success': self.COLOR_SUCCESS,
            'partial': self.COLOR_WARNING,
            'failed': self.COLOR_ERROR,
        }
        return status_colors.get(status, "DDDDDD")  # Default gray
